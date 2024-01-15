from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from json import load
from src.misc.tools.common import get_plot_path, get_solution_path


class Chart:
    def __init__(self, solution: str) -> None:
        self.solution = solution
        self.workbook = Workbook()
        self.out = get_plot_path(solution)
        self.source = get_solution_path(solution)
        self.best_sol_data = []
        self.accepted_sol_data = []
        self.levels = []

    def set_data(self) -> None:
        with open(self.source) as f:
            raw = load(f)
        if self.solution.lower() == "gda":
            self.accepted_sol_data = [
                [data["iter"], data["score"]] for data in raw["paths"]["accepted"]
            ]
            self.best_sol_data = [
                [data["iter"], data["score"]] for data in raw["paths"]["best_per_iter"]
            ]
            self.levels = [
                [data["iter"], data["level"]] for data in raw["paths"]["levels"]
            ]
        elif self.solution.lower() == "hsa":
            self.best_sol_data = [
                [data["iter"], data["score"]] for data in raw["paths"]["best_per_iter"]
            ]
        else:
            raise ValueError("Invalid solution set to Chart")

    def create_chart(self) -> None:
        headers = ["Iteration", "Distance"]
        data = self.best_sol_data.copy()
        data.insert(0, headers)
        wb = self.workbook
        best_sol_sheet = wb.active
        best_sol_sheet.title = f"{self.solution.upper()} Best Solutions for TSP" # type: ignore
        for row in data:
            best_sol_sheet.append(row)  # type: ignore
        
        c1 = LineChart()
        c1.title = f"{self.solution.upper()} Best Solutions for TSP"
        c1.style = 13
        c1.x_axis.title = "Iteration"
        c1.y_axis.title = "Distance"
        data = Reference(best_sol_sheet, min_col=2, min_row=1, max_col=3, max_row=len(data))
        c1.add_data(data, titles_from_data=True)

        s1 = c1.series[0]
        s1.graphicalProperties.line.solidFill = "00AAAA"
        s1.graphicalProperties.line.width = 9100

        best_sol_sheet.add_chart(c1, "A10") # type: ignore

        if self.solution.upper() == "GDA":
            accepted_sol_sheet = wb.create_sheet("GDA Accepted")
            wb.active = accepted_sol_sheet
            data = self.accepted_sol_data.copy()
            data.insert(0, headers)
            for row in data:
                accepted_sol_sheet.append(row)  # type: ignore
            
            c2 = LineChart()
            c2.title = f"{self.solution.upper()} Accepted Solutions for TSP"
            c2.style = 13
            c2.x_axis.title = "Iteration"
            c2.y_axis.title = "Distance"
            data = Reference(accepted_sol_sheet, min_col=2, min_row=1, max_col=3, max_row=len(data))
            c2.add_data(data, titles_from_data=True)

            s2 = c2.series[0]
            s2.graphicalProperties.line.solidFill = "00AAAA"
            s2.graphicalProperties.line.width = 9100

            accepted_sol_sheet.add_chart(c2, "A10") # type: ignore

            levels_sol_sheet = wb.create_sheet("GDA Levels")
            wb.active = levels_sol_sheet
            data = self.levels.copy()
            headers = ["Iteration", "Level"]
            data.insert(0, headers) 
            for row in data:
                levels_sol_sheet.append(row)  # type: ignore

            c3 = LineChart()
            c3.title = f"{self.solution.upper()} Levels for TSP"
            c3.style = 13
            c3.x_axis.title = "Iteration"
            c3.y_axis.title = "Level"
            data = Reference(levels_sol_sheet, min_col=2, min_row=1, max_col=3, max_row=len(data))
            c3.add_data(data, titles_from_data=True)

            s3 = c3.series[0]
            s3.graphicalProperties.line.solidFill = "00AAAA"
            s3.graphicalProperties.line.width = 9100

            levels_sol_sheet.add_chart(c3, "A10") # type: ignore

        wb.save(str(self.out))
